#!/usr/bin/env python3
"""
xlsx_to_csv.py — Convert the Vanchai product catalogue XLSX to CSV
===================================================================
Run this ONCE before running generate.py.

Usage:
    pip install openpyxl
    python xlsx_to_csv.py SEO-Automation-Product-catalogue.xlsx
"""

import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Install openpyxl: pip install openpyxl")

from config import CSV_FILE

COLUMNS = [
    "vendorArticleNumber", "vendorArticleName", "material",
    "productDetails", "productType", "amazonUrl", "myntraUrl",
    "nykaaUrl", "wixUrl", "imageUrl", "price", "category",
    "category_url", "json_ld",
]


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python xlsx_to_csv.py <path_to_xlsx>")

    xlsx_path = sys.argv[1]
    out_path  = Path(CSV_FILE)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Worksheet is empty.")

    header = [str(c).strip() if c else "" for c in rows[0]]
    print(f"Header columns found: {header}")

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
        writer.writeheader()

        skipped = 0
        written = 0
        seen    = set()

        for row_vals in rows[1:]:
            row = dict(zip(header, row_vals))

            def _str(val) -> str:
                """Coerce a cell value to string, returning '' for None/errors."""
                if val is None:
                    return ""
                s = str(val).strip()
                # Spreadsheet formula errors — treat as empty
                return "" if s.startswith("#") else s

            # CDN image domains that appear in URL columns but are NOT product pages
            _IMAGE_HOSTS = ("m.media-amazon.com", "images-amazon.com",
                            "static.wixstatic.com", "media.nykaa.com",
                            "assets.myntassets.com")

            def _url(val) -> str:
                """Clean a URL cell — strip errors, non-HTTP values, and image CDN links."""
                s = _str(val)
                if not s.startswith("http"):
                    return ""
                # Reject image CDN URLs (they are not product listing pages)
                from urllib.parse import urlparse
                host = urlparse(s).netloc.lower()
                if any(host == cdn or host.endswith("." + cdn) for cdn in _IMAGE_HOSTS):
                    return ""
                return s

            # Map to clean column names
            # Fields absent in this XLSX (material, productType, category_url, json_ld)
            # are left blank — generate.py handles missing fields gracefully.
            # extra_features and color are concatenated into productDetails for richer content.
            extra    = _str(row.get("extra_features"))
            color    = _str(row.get("color"))
            base_det = _str(row.get("productDetails"))
            details  = " | ".join(filter(None, [base_det, f"Color: {color}" if color else "", extra]))

            record = {
                "vendorArticleNumber": _str(row.get("vendorArticleNumber") or row.get("vendorArticleNumber\xa0")),
                "vendorArticleName":   _str(row.get("vendorArticleName")),
                "material":            _str(row.get("material")),
                "productDetails":      details,
                "productType":         _str(row.get("productType")),
                "amazonUrl":           _url(row.get("amazonUrl")),
                "myntraUrl":           _url(row.get("myntraUrl")),
                "nykaaUrl":            _url(row.get("nykaaUrl")),
                "wixUrl":              _url(row.get(" wixUrl") or row.get("wixUrl")),
                "imageUrl":            _url(row.get("imageUrl")),
                "price":               _str(row.get("price")),
                "category":            _str(row.get("category")),
                "category_url":        _url(row.get("category_url")),
                "json_ld":             "",  # built from scratch by generate.py
            }

            name = record["vendorArticleName"]
            sku  = record["vendorArticleNumber"]

            # Skip broken rows
            if not name or name.startswith("http") or "Not Found" in name or name.startswith("Key ["):
                skipped += 1
                continue

            # Require at least one valid platform URL to send traffic somewhere
            if not any([record["wixUrl"], record["amazonUrl"], record["myntraUrl"], record["nykaaUrl"]]):
                skipped += 1
                continue

            # Deduplicate by sku+name
            key = f"{sku}|{name}"
            if key in seen:
                skipped += 1
                continue
            seen.add(key)

            writer.writerow(record)
            written += 1

    print(f"\nDone. {written} products written to {out_path}  ({skipped} rows skipped).")


if __name__ == "__main__":
    main()
